import math
import pathlib
import platform
import random
import string
import threading
import time
from io import BytesIO
from itertools import product
from urllib.parse import urljoin

import execjs
import progressbar
import requests
from imgcat import imgcat
from openpyxl import Workbook
from openpyxl.drawing.image import Image as openpyxlImage
from PIL import Image
from pyecharts import options as opts
from pyecharts.charts import Sunburst

from webot.common import (
    check_if_can_open,
    error_log,
    format_sunburst_city,
    get_pic,
    random_color,
)
from webot.conf import conf
from webot.data import (
    API_analysis_path,
    API_conf_path,
    API_media_icon_path,
    API_media_path,
    API_target,
)
from webot.log import debug, success, warning


class Device:
    @staticmethod
    @error_log()
    def create_device_id() -> str:
        """
            创建随机设备指纹
        """
        device_id = f"e{str(random.random())[2:17]}"
        debug(f"device_id: {device_id}")
        return device_id

    @staticmethod
    @error_log()
    def create_client_msg_id() -> str:
        """
            创建客户端信息指纹
        """
        client_msg_id = int(time.time() * 1e3) + random.randint(1, 1000)
        debug(f"client_msg_id: {client_msg_id}")
        return client_msg_id

    @staticmethod
    @error_log()
    def get_timestamp(reverse: bool = False) -> int:
        """
            获取时间戳,支持取反
        """
        timestamp = None
        if not reverse:
            timestamp = int(time.time() * 1e3)
        else:
            timestamp = execjs.eval("~new Date")
        debug(f"timestamp: {timestamp}")
        return timestamp

    @staticmethod
    @error_log(raise_exit=True)
    def show_qrcode(buffer, local_func) -> None:
        """
            打印图片
        """
        if platform.system().lower() != "darwin":
            imgcat(buffer)
        else:
            code_img = threading.Thread(target=local_func, args=(buffer,))
            code_img.start()

    @error_log()
    def trans_map(contacts: dict, batch_contacts: dict) -> dict:
        """
            创建名称列表
        """
        choice_name = (
            lambda item: item.get("RemarkName", "").strip()
            if item.get("RemarkName", "").strip()
            else item.get("NickName", "").strip()
        )
        person_map = {
            item["UserName"]: choice_name(item) for item in contacts["MemberList"]
        }
        for _ in batch_contacts["ContactList"]:
            person_map[_["UserName"]] = _["NickName"]
            for _ in _["MemberList"]:
                person_map[_["UserName"]] = _["NickName"]
        return person_map

    @staticmethod
    def filters(types: bool = None, is_me: bool = False, is_group: bool = False):
        """
            消息过滤
        """

        def decorator(func):
            def wrapper(obj, msg, *args, **kwargs):
                nonlocal types
                if not types or not isinstance(types, list):
                    types = []
                if msg["type"] not in types:
                    return
                if is_group:
                    if not "@@" in msg["from"]:
                        return
                if is_me:
                    if not msg["from"] == conf.my_id:
                        return
                return func(obj, msg, *args, **kwargs)

            return wrapper

        return decorator

    @error_log()
    def export_all_contact(
        contacts: dict, session: requests.Session, person_data: dict
    ) -> pathlib.Path:
        """
            导出通讯录
        """
        warning("Contact exporting...")
        wb = Workbook()
        sheet = wb.active
        keys = contacts["MemberList"][0].keys()
        for x, key in enumerate(keys):
            sheet.cell(row=1, column=x + 1, value=key)
        for y, item in progressbar.progressbar(enumerate(contacts["MemberList"])):
            y = y + 2
            for x, key in enumerate(keys):
                x = x + 1
                data = item[key]

                if key != "HeadImgUrl":
                    if key == "MemberList":
                        data = "".join(data)
                    sheet.cell(row=y, column=x, value=data)
                else:
                    pic = get_pic(
                        session,
                        urljoin(API_target, data),
                        API_media_icon_path
                        / f"{item['PYInitial']}_{item['VerifyFlag']}.png",
                    )
                    if pic:
                        x = x - 1
                        index_code = string.ascii_uppercase[x]
                        size = (50, 50)
                        sheet.column_dimensions[index_code].width, sheet.row_dimensions[
                            y
                        ].height = size
                        img = openpyxlImage(BytesIO(pic))
                        img.width, img.height = size
                        sheet.add_image(img, f"{index_code}{y}")
                    else:
                        sheet.cell(row=y, column=x, value="")
        path = API_analysis_path / f'{person_data["User"]["NickName"]}_contacts.xlsx'
        wb.save(path)
        success(f"export contacts: {path}")
        return path

    @error_log()
    def export_sunburst_city(
        data: dict, image_result_save_path: str = "./sunburst_city.html"
    ) -> pathlib.Path:
        """
            导出城市分布图
        """
        image = (
            Sunburst(init_opts=opts.InitOpts(width="1000px", height="600px"))  # 设定画布长宽
            .add(
                "",
                data_pair=format_sunburst_city(data),  # 载入数据
                highlight_policy="ancestor",
                radius=[0, "95%"],
                sort_="null",
                levels=[
                    {},  # 第一圈样式，如果有国家的话就不会空着
                    {
                        "r0": "15%",
                        "r": "45%",
                        "itemStyle": {"borderWidth": 2},
                        "label": {"rotate": "tangential"},
                    },  # 第二圈样式，对标省
                    {
                        "r0": "35%",
                        "r": "70%",
                        "label": {"position": "outside", "padding": 3, "silent": False},
                        "itemStyle": {"borderWidth": 1},
                    },  # 最外圈样式，对标市
                ],
            )
            # 设定标题
            .set_global_opts(title_opts=opts.TitleOpts(title="Sunburst-城市分布"))
            .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}"))  # 设定名称
        )
        path = image.render(image_result_save_path)
        success(f"make city sunburst: {path}")
        return pathlib.Path(path)

    @error_log()
    def make_icon_wall(
        image_root_path: str,
        image_result_save_path: str = "./icon_wall.png",
        length: int = 1440,
        weight: int = 900,
        patterns: str = "*",
    ) -> pathlib.Path:
        """
            生成图片墙
            patterns: *_0 好友
                  *   全部
        """
        images = list(
            filter(
                check_if_can_open,
                pathlib.Path(image_root_path).glob(f"**/{patterns}.png"),
            )
        )
        per_size = int(math.sqrt(length * weight / len(images)))
        image = Image.new("RGBA", (length, weight))
        for indexs, (x, y) in enumerate(
            product(range(int(length / per_size)), range(int(weight / per_size)))
        ):
            img = Image.open(images[indexs])
            img = img.resize((per_size, per_size), Image.ANTIALIAS)
            image.paste(img, (x * per_size, y * per_size))
        image.save(image_result_save_path)
        success(f"make icon wall: {image_result_save_path}")
        return pathlib.Path(image_result_save_path)
